import openpyxl
import os

d = r"D:\App_Claude_Antigravity\Bao-gia\MAU-BAO-GIA"
out_path = r"D:\App_Claude_Antigravity\Bao-gia\exports\inspect_excel.txt"

os.makedirs(os.path.dirname(out_path), exist_ok=True)

with open(out_path, "w", encoding="utf-8") as out:
    for filename in os.listdir(d):
        if not filename.endswith(".xlsx"):
            continue
        
        filepath = os.path.join(d, filename)
        out.write(f"=== FILE: {filename} ===\n")
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            for sheetname in wb.sheetnames:
                out.write(f"  Sheet: {sheetname}\n")
                ws = wb[sheetname]
                # Read first 30 rows and 10 columns
                for r in range(1, 40):
                    row_vals = []
                    for c in range(1, 15):
                        val = ws.cell(row=r, column=c).value
                        row_vals.append(val)
                    # Only write if there's some content in the row
                    if any(v is not None for v in row_vals):
                        row_str = " | ".join(str(v) if v is not None else "" for v in row_vals)
                        out.write(f"    Row {r:02d}: {row_str}\n")
                out.write("\n")
        except Exception as e:
            out.write(f"  Error reading file: {e}\n\n")

print("Inspection completed!")
