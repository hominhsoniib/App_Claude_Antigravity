import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from models.models import Product, Customer

PRODUCT_TEMPLATE_COLUMNS = [
    "sku", "name", "spec", "unit", "cost_price", "list_price",
    "max_discount_pct", "vat_pct", "category", "warehouse", "stock_qty"
]

CUSTOMER_TEMPLATE_COLUMNS = [
    "code", "name", "tax_code", "address", "group", "region"
]


def build_template(columns, sheet_name="Template"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def validate_products_df(df):
    errors = []
    required = ["sku", "name", "unit", "list_price"]
    for col in required:
        if col not in df.columns:
            errors.append(f"Thiếu cột bắt buộc: {col}")
    if errors:
        return errors
    for i, row in df.iterrows():
        if pd.isna(row.get("sku")) or pd.isna(row.get("name")):
            errors.append(f"Dòng {i+2}: thiếu mã hàng hoặc tên hàng")
        if pd.isna(row.get("list_price")) or float(row.get("list_price", 0)) <= 0:
            errors.append(f"Dòng {i+2}: giá bán không hợp lệ")
    return errors


def import_products(db, df, actor="system"):
    """Import hàng loạt sản phẩm, trả về (số dòng thành công, số dòng lỗi, danh sách lỗi)."""
    errors = validate_products_df(df)
    if errors:
        return 0, len(df), errors

    success = 0
    for _, row in df.iterrows():
        existing = db.query(Product).filter(Product.sku == row["sku"]).first()
        values = dict(
            sku=row["sku"], name=row["name"],
            spec=row.get("spec", ""), unit=row.get("unit", ""),
            cost_price=float(row.get("cost_price", 0) or 0),
            list_price=float(row.get("list_price", 0) or 0),
            max_discount_pct=float(row.get("max_discount_pct", 0) or 0),
            vat_pct=float(row.get("vat_pct", 10) or 10),
            category=row.get("category", ""), warehouse=row.get("warehouse", ""),
            stock_qty=float(row.get("stock_qty", 0) or 0),
        )
        if existing:
            for k, v in values.items():
                setattr(existing, k, v)
        else:
            db.add(Product(**values))
        success += 1
    db.commit()
    return success, 0, []


def validate_customers_df(df):
    errors = []
    required = ["code", "name"]
    for col in required:
        if col not in df.columns:
            errors.append(f"Thiếu cột bắt buộc: {col}")
    if errors:
        return errors
    for i, row in df.iterrows():
        if pd.isna(row.get("code")) or pd.isna(row.get("name")):
            errors.append(f"Dòng {i+2}: thiếu mã khách hàng hoặc tên khách hàng")
    return errors


def import_customers(db, df, actor="system"):
    """Import hàng loạt khách hàng, trả về (số dòng thành công, số dòng lỗi, danh sách lỗi)."""
    errors = validate_customers_df(df)
    if errors:
        return 0, len(df), errors

    success = 0
    for _, row in df.iterrows():
        existing = db.query(Customer).filter(Customer.code == row["code"]).first()
        values = dict(
            code=str(row["code"]).strip(),
            name=str(row["name"]).strip(),
            tax_code=str(row.get("tax_code", "") or "").strip(),
            address=str(row.get("address", "") or "").strip(),
            group=str(row.get("group", "") or "").strip(),
            region=str(row.get("region", "") or "").strip(),
        )
        
        # Avoid putting nan values into database
        for k, v in values.items():
            if pd.isna(v) or v.lower() == "nan":
                values[k] = ""
                
        if existing:
            for k, v in values.items():
                setattr(existing, k, v)
        else:
            db.add(Customer(**values))
        success += 1
    db.commit()
    return success, 0, []


def export_quotation_excel(header):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bao gia"
    ws.append(["BÁO GIÁ", header.quote_no])
    ws.append(["Khách hàng", header.customer.name if header.customer else ""])
    ws.append(["Ngày báo giá", header.quote_date.strftime("%d/%m/%Y")])
    ws.append([])
    ws.append(["STT", "Sản phẩm", "ĐVT", "SL", "Đơn giá", "CK%", "VAT%", "Thành tiền"])
    for i, d in enumerate(header.details, start=1):
        ws.append([i, d.product.name if d.product else "", d.product.unit if d.product else "",
                   d.qty, d.unit_price, d.discount_pct, d.vat_pct, d.line_total])
    ws.append([])
    ws.append(["", "", "", "", "", "", "Tạm tính", header.subtotal])
    ws.append(["", "", "", "", "", "", "VAT", header.vat_total])
    ws.append(["", "", "", "", "", "", "TỔNG CỘNG", header.grand_total])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_omri_quotation_excel(file_content, sheet_name):
    """
    Phân tích file Excel báo giá mẫu OMRI của khách hàng tải lên.
    Trả về danh sách các dòng sản phẩm gồm: {"name", "spec", "unit", "price"}.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Không tìm thấy sheet '{sheet_name}' trong file Excel.")
    
    ws = wb[sheet_name]
    header_row_idx = None
    col_map = {}
    
    # 1. Quét tìm dòng tiêu đề chứa tên sản phẩm
    for r in range(1, 40):
        row_vals = [str(ws.cell(row=r, column=c).value or "").strip().lower() for c in range(1, 20)]
        has_name = any("tên hàng" in val or "tên sản phẩm" in val for val in row_vals)
        has_price = any("giá" in val or "đơn giá" in val for val in row_vals)
        
        if has_name and has_price:
            header_row_idx = r
            # Map các cột
            for c in range(1, 20):
                val = str(ws.cell(row=r, column=c).value or "").strip().lower()
                if "tên hàng" in val or "tên sản phẩm" in val:
                    col_map["name"] = c
                elif "quy cách" in val or "thể tích" in val:
                    col_map["spec"] = c
                elif "đvt" in val or "loại" in val or "đơn vị tính" in val:
                    col_map["unit"] = c
                elif "giá" in val or "đơn giá" in val:
                    # Nếu có cả "đơn giá" và "giá thùng", ưu tiên cột "đơn giá"
                    if "thùng" in val and "price" in col_map:
                        pass
                    else:
                        col_map["price"] = c
            break

    if header_row_idx is None:
        raise ValueError("Không thể nhận diện cấu trúc bảng báo giá (thiếu cột Tên hàng hóa hoặc Đơn giá).")

    if "name" not in col_map or "price" not in col_map:
        raise ValueError("File Excel thiếu cột bắt buộc: Tên hàng hóa hoặc Đơn giá.")

    items = []
    # 2. Đọc các dòng sản phẩm tiếp theo dưới dòng tiêu đề
    for r in range(header_row_idx + 1, 150):
        name_val = ws.cell(row=r, column=col_map["name"]).value
        price_val = ws.cell(row=r, column=col_map["price"]).value
        
        if not name_val:
            # Dừng nếu 3 dòng liên tiếp trống hoàn toàn
            empty_count = 0
            for offset in range(3):
                if not ws.cell(row=r+offset, column=col_map["name"]).value:
                    empty_count += 1
            if empty_count == 3:
                break
            continue
            
        name_str = str(name_val).strip()
        if any(stop_word in name_str.lower() for stop_word in ["công ty", "trân trọng", "giám đốc", "tổng cộng", "chân thành", "ghi chú", "ngày"]):
            break
            
        try:
            if price_val is None:
                continue
            # Làm sạch chuỗi giá trị số
            price_clean = str(price_val).replace(",", "").replace("đ", "").strip()
            price = float(price_clean)
            if price <= 0:
                continue
        except (ValueError, TypeError):
            continue
            
        spec = ""
        if "spec" in col_map:
            spec = str(ws.cell(row=r, column=col_map["spec"]).value or "").strip()
            
        unit = "Cái"
        if "unit" in col_map:
            unit = str(ws.cell(row=r, column=col_map["unit"]).value or "").strip()
            if not unit:
                unit = "Cái"
                
        # Làm sạch khoảng trắng dư thừa
        clean_name = " ".join(name_str.split())
        
        items.append({
            "name": clean_name,
            "spec": spec,
            "unit": unit,
            "price": price
        })
        
    return items


def validate_opening_balance_df(df):
    errors = []
    required = ["sku", "stock_qty"]
    for col in required:
        if col not in df.columns:
            errors.append(f"Thiếu cột bắt buộc: {col}")
    if errors:
        return errors
    for i, row in df.iterrows():
        if pd.isna(row.get("sku")):
            errors.append(f"Dòng {i+2}: Thiếu cột mã hàng (sku)")
        if pd.isna(row.get("stock_qty")):
            errors.append(f"Dòng {i+2}: Thiếu cột số lượng tồn đầu kỳ (stock_qty)")
        else:
            try:
                val = float(row.get("stock_qty"))
                if val < 0:
                    errors.append(f"Dòng {i+2}: Số lượng tồn đầu kỳ không được âm")
            except (ValueError, TypeError):
                errors.append(f"Dòng {i+2}: Số lượng tồn đầu kỳ '{row.get('stock_qty')}' không hợp lệ (phải là số)")
    return errors


def build_opening_balance_template(db):
    """Xây dựng file Excel mẫu số dư đầu kỳ, tự điền sẵn danh sách sản phẩm hiện tại."""
    products = db.query(Product).order_by(Product.sku.asc()).all()
    rows = []
    for p in products:
        rows.append({
            "sku": p.sku,
            "product_name": p.name,
            "stock_qty": p.stock_qty or 0.0,
            "note": "Khai báo số dư đầu kỳ"
        })
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Opening_Balances", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Opening_Balances"]
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        for col_idx in range(1, 5):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            worksheet.column_dimensions[cell.column_letter].width = 25
    buf.seek(0)
    return buf


def import_opening_balances(db, df, actor="system"):
    """Import số dư đầu kỳ hàng loạt từ Excel."""
    from services.inventory_service import add_transaction
    
    errors = validate_opening_balance_df(df)
    if errors:
        return 0, len(df), errors
        
    success = 0
    fail = 0
    errs = []
    
    for i, row in df.iterrows():
        sku = str(row["sku"]).strip()
        qty = row["stock_qty"]
        note = str(row.get("note", "") or "").strip()
        if pd.isna(row.get("note")) or note.lower() == "nan" or not note:
            note = "Khai báo số dư đầu kỳ (Import Excel)"
            
        product = db.query(Product).filter(Product.sku == sku).first()
        if not product:
            fail += 1
            errs.append(f"Dòng {i+2}: Mã sản phẩm '{sku}' không tồn tại trong hệ thống.")
            continue
            
        try:
            qty_val = float(qty)
            if qty_val < 0:
                fail += 1
                errs.append(f"Dòng {i+2}: Số lượng tồn đầu kỳ '{qty}' không được âm.")
                continue
                
            add_transaction(
                db=db,
                product_id=product.id,
                transaction_type="OPENING",
                qty=qty_val,
                reference_no="OPENING",
                note=note,
                actor=actor
            )
            success += 1
        except Exception as e:
            fail += 1
            errs.append(f"Dòng {i+2}: Lỗi khi xử lý dòng hàng '{sku}': {str(e)}")
            
    return success, fail, errs


def build_quotation_import_template(db):
    """Xây dựng file Excel mẫu tạo báo giá, tự điền sẵn sản phẩm để người dùng nhập số lượng/chiết khấu."""
    products = db.query(Product).order_by(Product.sku.asc()).all()
    rows = []
    for p in products:
        rows.append({
            "sku": p.sku,
            "product_name": p.name,
            "qty": 0.0,
            "unit_price": p.list_price or 0.0,
            "discount_pct": 0.0,
            "vat_pct": p.vat_pct or 10.0,
            "line_note": ""
        })
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Quotation_Template", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Quotation_Template"]
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        for col_idx in range(1, 8):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            worksheet.column_dimensions[cell.column_letter].width = 22
    buf.seek(0)
    return buf


def validate_quotation_import_df(df):
    errors = []
    required = ["sku", "qty", "unit_price", "discount_pct", "vat_pct"]
    for col in required:
        if col not in df.columns:
            errors.append(f"Thiếu cột bắt buộc trong file: {col}")
    return errors


def import_quotation_from_excel(db, df, customer_id, sales_rep_id, actor="system"):
    """Tạo báo giá mới từ file Excel mẫu chuẩn."""
    from services.quotation_service import create_quotation, add_line
    
    errors = validate_quotation_import_df(df)
    if errors:
        return 0, len(df), errors, None
        
    # Lọc các dòng có số lượng > 0
    valid_rows = []
    for idx, row in df.iterrows():
        try:
            qty_val = float(row["qty"])
            if qty_val > 0:
                valid_rows.append((idx, row, qty_val))
        except (ValueError, TypeError):
            continue
            
    if not valid_rows:
        return 0, 0, ["Không tìm thấy dòng sản phẩm nào có Số lượng (qty) lớn hơn 0 trong file Excel."], None
        
    # Tạo báo giá header
    header = create_quotation(
        db=db,
        customer_id=customer_id,
        sales_rep_id=sales_rep_id,
        valid_days=15,
        payment_terms="Thanh toán 50% đặt cọc, 50% trước khi nhận hàng",
        delivery_terms="Giao hàng tại kho khách hàng",
        lead_time="7 ngày làm việc",
        note=f"Khởi tạo từ file Excel mẫu chuẩn ngày {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        actor=actor
    )
    
    success = 0
    fail = 0
    errs = []
    
    for idx, row, qty_val in valid_rows:
        sku = str(row["sku"]).strip()
        product = db.query(Product).filter(Product.sku == sku).first()
        if not product:
            fail += 1
            errs.append(f"Dòng {idx+2}: Mã sản phẩm '{sku}' không tồn tại.")
            continue
            
        try:
            price_val = float(row["unit_price"])
            discount_val = float(row.get("discount_pct", 0.0) or 0.0)
            vat_val = float(row.get("vat_pct", 10.0) or 10.0)
            
            add_line(
                db=db,
                header_id=header.id,
                product_id=product.id,
                qty=qty_val,
                unit_price=price_val,
                discount_pct=discount_val,
                vat_pct=vat_val
            )
            success += 1
        except Exception as e:
            fail += 1
            errs.append(f"Dòng {idx+2}: Lỗi khi thêm sản phẩm '{sku}': {str(e)}")
            
    return success, fail, errs, header.quote_no


